"""
GET  /api/v1/results/{sinav_id}            — sonuç listesi
GET  /api/v1/results/{sinav_id}/statistics — istatistikler
GET  /api/v1/results/{sinav_id}/export     — Excel export
"""
from __future__ import annotations

import io
import statistics
from collections import defaultdict
from difflib import SequenceMatcher

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from middleware.auth_middleware import verify_firebase_token
from models.schemas import IstatistikResponse, SonucListResponse
from services import firebase_service as fb

router = APIRouter(prefix="/api/v1/results", tags=["Sonuçlar"])

# Renk filleri
TURUNCU = PatternFill("solid", fgColor="F4B942")

# ─────────────────────────── Eşleştirme Yardımcıları ─────────────────

def _normalize(s: str) -> str:
    return (
        s.upper().strip()
        .replace("İ", "I").replace("Ğ", "G").replace("Ü", "U")
        .replace("Ş", "S").replace("Ö", "O").replace("Ç", "C")
    )


def _benzerlik(a: str, b: str) -> float:
    return SequenceMatcher(None, _normalize(a), _normalize(b)).ratio()


def _esles(ogrenci_no: str, ad_soyad: str, liste: list[dict]) -> dict | None:
    """
    Önce öğrenci numarasıyla, bulamazsa isim benzerliğiyle (%85+) eşleştirir.
    liste: [{"no": str, "ad": str, "soyad": str, "satir": int}, ...]
    """
    no_temiz = ogrenci_no.strip().lstrip("0") if ogrenci_no != "?????????" else None

    # 1. Numara eşleştirme
    if no_temiz:
        for ogr in liste:
            if ogr["no"].strip().lstrip("0") == no_temiz:
                return ogr

    # 2. Tam isim benzerliği
    for ogr in liste:
        tam_ad = f"{ogr['ad']} {ogr['soyad']}"
        if _benzerlik(ad_soyad, tam_ad) >= 0.85:
            return ogr

    return None
YESIL = PatternFill("solid", fgColor="C6EFCE")
KIRMIZI = PatternFill("solid", fgColor="FFC7CE")
GRI = PatternFill("solid", fgColor="D9D9D9")
MAVI = PatternFill("solid", fgColor="BDD7EE")
SARI = PatternFill("solid", fgColor="FFEB9C")


def _sinav_yetki_kontrol(sinav_id: str, uid: str) -> dict:
    sinav = fb.sinav_getir(sinav_id)
    if not sinav:
        raise HTTPException(status_code=404, detail="Sınav bulunamadı")
    if sinav.get("ogretmen_id") != uid:
        raise HTTPException(status_code=403, detail="Bu sınava erişim yetkiniz yok")
    return sinav


@router.get("/{sinav_id}", response_model=SonucListResponse, summary="Sınav sonuçları")
async def get_results(
    sinav_id: str,
    token_data: dict = Depends(verify_firebase_token),
):
    uid = token_data["uid"]
    _sinav_yetki_kontrol(sinav_id, uid)
    sonuclar = fb.sinav_sonuclari(sinav_id)
    return SonucListResponse(sonuclar=sonuclar, toplam=len(sonuclar))


@router.get("/{sinav_id}/statistics", response_model=IstatistikResponse, summary="Sınav istatistikleri")
async def get_statistics(
    sinav_id: str,
    token_data: dict = Depends(verify_firebase_token),
):
    uid = token_data["uid"]
    sinav = _sinav_yetki_kontrol(sinav_id, uid)
    sonuclar = fb.sinav_sonuclari(sinav_id)

    if not sonuclar:
        raise HTTPException(status_code=404, detail="Bu sınava ait sonuç bulunamadı")

    puanlar = [float(s.get("puan", 0)) for s in sonuclar]
    soru_sayisi = sinav.get("soru_sayisi", 20)
    cevap_anahtari: dict[str, str] = sinav.get("cevap_anahtari", {})

    # Soru bazlı başarı oranı
    soru_dogru: dict[str, int] = defaultdict(int)
    for s in sonuclar:
        cevaplar = s.get("cevaplar", {})
        for soru_no, dogru_cevap in cevap_anahtari.items():
            if str(cevaplar.get(soru_no, "")).upper() == dogru_cevap.upper():
                soru_dogru[soru_no] += 1

    n = len(sonuclar)
    basari_oranlari = {k: round(v / n, 3) for k, v in soru_dogru.items()}

    return IstatistikResponse(
        sinav_id=sinav_id,
        ogrenci_sayisi=n,
        ortalama=round(statistics.mean(puanlar), 2),
        medyan=round(statistics.median(puanlar), 2),
        min_puan=min(puanlar),
        max_puan=max(puanlar),
        standart_sapma=round(statistics.stdev(puanlar) if n > 1 else 0.0, 2),
        soru_basari_oranlari=basari_oranlari,
    )


@router.get("/{sinav_id}/export", summary="Excel export")
async def export_excel(
    sinav_id: str,
    token_data: dict = Depends(verify_firebase_token),
):
    uid = token_data["uid"]
    sinav = _sinav_yetki_kontrol(sinav_id, uid)
    sonuclar = fb.sinav_sonuclari(sinav_id)

    soru_sayisi: int = sinav.get("soru_sayisi", 20)
    cevap_anahtari: dict[str, str] = sinav.get("cevap_anahtari", {})

    wb = openpyxl.Workbook()

    # ── Sayfa 1: Özet ─────────────────────────────────────────────────
    ozet = wb.active
    ozet.title = "Özet"

    basliklar = ["Sıra", "Ad Soyad", "Öğrenci No", "Doğru", "Yanlış", "Boş", "Puan", "Durum"]
    for col, baslik in enumerate(basliklar, 1):
        h = ozet.cell(1, col, baslik)
        h.font = Font(bold=True)
        h.fill = MAVI
        h.alignment = Alignment(horizontal="center")

    for row_i, s in enumerate(sonuclar, 2):
        puan = float(s.get("puan", 0))
        guvenskor = float(s.get("guvenskor", 1.0))
        durum = "✓" if guvenskor >= 0.85 else "⚠ Manuel Kontrol"

        degerler = [
            row_i - 1,
            s.get("ad_soyad", "?"),
            s.get("ogrenci_no", "?"),
            s.get("dogru", 0),
            s.get("yanlis", 0),
            s.get("bos", 0),
            puan,
            durum,
        ]
        fill = YESIL if guvenskor >= 0.85 else SARI

        for col_i, deger in enumerate(degerler, 1):
            c = ozet.cell(row_i, col_i, deger)
            c.fill = fill
            c.alignment = Alignment(horizontal="center")

    # Sütun genişlikleri
    for i, gen in enumerate([6, 25, 15, 8, 8, 8, 10, 20], 1):
        ozet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = gen

    # ── Sayfa 2: Detay ────────────────────────────────────────────────
    detay = wb.create_sheet("Detay")

    # Başlık satırı
    detay.cell(1, 1, "Ad Soyad").font = Font(bold=True)
    detay.cell(1, 2, "Öğrenci No").font = Font(bold=True)
    for q in range(1, soru_sayisi + 1):
        c = detay.cell(1, q + 2, f"S{q}")
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")

    # Cevap anahtarı satırı (mavi)
    detay.cell(2, 1, "CEVAP ANAHTARI").font = Font(bold=True)
    detay.cell(2, 1).fill = MAVI
    detay.cell(2, 2, "-").fill = MAVI
    for q in range(1, soru_sayisi + 1):
        c = detay.cell(2, q + 2, cevap_anahtari.get(str(q), "?"))
        c.fill = MAVI
        c.alignment = Alignment(horizontal="center")

    # Öğrenci satırları
    for row_i, s in enumerate(sonuclar, 3):
        detay.cell(row_i, 1, s.get("ad_soyad", "?"))
        detay.cell(row_i, 2, s.get("ogrenci_no", "?"))
        cevaplar = s.get("cevaplar", {})

        for q in range(1, soru_sayisi + 1):
            ogr_cevap = cevaplar.get(str(q), "BOS")
            dogru_cevap = cevap_anahtari.get(str(q), "")
            c = detay.cell(row_i, q + 2, ogr_cevap)
            c.alignment = Alignment(horizontal="center")
            if ogr_cevap == "BOS":
                c.fill = GRI
            elif ogr_cevap.upper() == dogru_cevap.upper():
                c.fill = YESIL
            else:
                c.fill = KIRMIZI

    # Excel dosyasını bellekte oluştur
    excel_buf = io.BytesIO()
    wb.save(excel_buf)
    excel_buf.seek(0)

    sinav_adi = sinav.get("ad", sinav_id).replace(" ", "_")
    dosya_adi = f"sonuclar_{sinav_adi}.xlsx"

    return StreamingResponse(
        excel_buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{dosya_adi}"'},
    )


@router.post("/{sinav_id}/export-with-list", summary="Öğrenci listesine not yaz")
async def export_with_student_list(
    sinav_id: str,
    ogrenci_listesi: UploadFile = File(..., description="Öğrenci listesi Excel (.xlsx)"),
    token_data: dict = Depends(verify_firebase_token),
):
    """
    Yüklenen öğrenci listesi Excel'ine tarama sonuçlarını eşleştirerek Notlar sütununu doldurur.
    Eşleşme önceliği: 1) Öğrenci No  2) Ad Soyad benzerliği (%85+)
    Listede olmayan öğrenciler ve okunamayan kağıtlar boş bırakılır.

    Beklenen Excel sütunları: Öğrenci No | Öğrenci Adı | Öğrenci Soyadı | Notlar
    """
    uid = token_data["uid"]
    _sinav_yetki_kontrol(sinav_id, uid)
    sonuclar = fb.sinav_sonuclari(sinav_id)

    # ── Öğrenci listesi Excel'ini oku ────────────────────────────────
    icerik = await ogrenci_listesi.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(icerik))
    except Exception:
        raise HTTPException(status_code=400, detail="Excel dosyası okunamadı. Geçerli bir .xlsx dosyası yükleyin.")

    ws = wb.active

    # Başlık satırını bul (Öğrenci No sütununu ara)
    baslik_satir = None
    sutunlar: dict[str, int] = {}
    for row in ws.iter_rows(min_row=1, max_row=10):
        for cell in row:
            deger = str(cell.value or "").strip().lower()
            if "öğrenci no" in deger or "ogrenci no" in deger:
                baslik_satir = cell.row
                break
        if baslik_satir:
            break

    if not baslik_satir:
        raise HTTPException(status_code=400, detail="Excel'de 'Öğrenci No' başlık sütunu bulunamadı.")

    # Sütun indekslerini belirle
    for cell in ws[baslik_satir]:
        deger = str(cell.value or "").strip().lower()
        if "öğrenci no" in deger or "ogrenci no" in deger:
            sutunlar["no"] = cell.column
        elif "adı" in deger or "adi" in deger:
            sutunlar["ad"] = cell.column
        elif "soyadı" in deger or "soyadi" in deger:
            sutunlar["soyad"] = cell.column
        elif "not" in deger:
            sutunlar["not"] = cell.column

    # Notlar sütunu yoksa oluştur
    if "not" not in sutunlar:
        yeni_sutun = ws.max_column + 1
        ws.cell(baslik_satir, yeni_sutun, "Notlar").font = Font(bold=True)
        ws.cell(baslik_satir, yeni_sutun).fill = MAVI
        ws.cell(baslik_satir, yeni_sutun).alignment = Alignment(horizontal="center")
        sutunlar["not"] = yeni_sutun

    # Öğrenci listesini yükle
    liste: list[dict] = []
    for row in ws.iter_rows(min_row=baslik_satir + 1):
        no_val = str(ws.cell(row[0].row, sutunlar.get("no", 1)).value or "").strip()
        ad_val = str(ws.cell(row[0].row, sutunlar.get("ad", 2)).value or "").strip()
        soyad_val = str(ws.cell(row[0].row, sutunlar.get("soyad", 3)).value or "").strip()
        if no_val or ad_val:
            liste.append({
                "no": no_val,
                "ad": ad_val,
                "soyad": soyad_val,
                "satir": row[0].row,
            })

    # ── Notları eşleştir ve yaz ──────────────────────────────────────
    eslesmeyenler: list[str] = []

    for s in sonuclar:
        ogrenci_no = str(s.get("ogrenci_no", "?????????"))
        ad_soyad = str(s.get("ad_soyad", "?"))
        puan = s.get("puan", "")
        guvenskor = float(s.get("guvenskor", 1.0))

        # Okunamayan kağıt → atla
        if ogrenci_no == "?????????" and ad_soyad in ("?", ""):
            continue

        eslesme = _esles(ogrenci_no, ad_soyad, liste)

        if eslesme:
            not_hucresi = ws.cell(eslesme["satir"], sutunlar["not"])
            not_hucresi.value = puan
            not_hucresi.alignment = Alignment(horizontal="center")
            # Düşük güven skorunda uyarı rengi
            not_hucresi.fill = SARI if guvenskor < 0.70 else YESIL
        else:
            eslesmeyenler.append(f"{ogrenci_no} - {ad_soyad}")

    # ── Eşleşmeyen uyarı sayfası ─────────────────────────────────────
    if eslesmeyenler:
        uyari_ws = wb.create_sheet("Eşleşmeyen Sonuçlar")
        uyari_ws.cell(1, 1, "Listede Bulunamayan Tarama Sonuçları").font = Font(bold=True)
        uyari_ws.cell(1, 1).fill = KIRMIZI
        uyari_ws.cell(2, 1, "Öğrenci No - Ad Soyad").font = Font(bold=True)
        for i, satir in enumerate(eslesmeyenler, 3):
            uyari_ws.cell(i, 1, satir)
        uyari_ws.column_dimensions["A"].width = 40

    # ── Dosyayı döndür ───────────────────────────────────────────────
    excel_buf = io.BytesIO()
    wb.save(excel_buf)
    excel_buf.seek(0)

    sinav_adi = sinav_id.replace(" ", "_")
    dosya_adi = f"notlar_{sinav_adi}.xlsx"

    return StreamingResponse(
        excel_buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{dosya_adi}"'},
    )
