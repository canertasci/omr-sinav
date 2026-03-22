"""
POST /api/v1/scan/single       — tek kağıt okuma
POST /api/v1/scan/batch        — toplu kağıt okuma (max 30)
POST /api/v1/scan/excel-sinav  — Excel listesiyle toplu tarama + Excel çıktısı
"""
from __future__ import annotations

import base64
import io
import os
from concurrent.futures import ThreadPoolExecutor, as_completed

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from fastapi import APIRouter, Depends, HTTPException, Request
from slowapi import Limiter
from slowapi.util import get_remote_address

from middleware.auth_middleware import verify_firebase_token
from models.schemas import (
    BatchSonuc,
    ExcelSinavRequest,
    ExcelSinavResponse,
    KontrolGerekli,
    ScanBatchRequest,
    ScanBatchResponse,
    ScanResponse,
    ScanSingleRequest,
)
from services import firebase_service as fb
from services.omr_engine import kagit_oku
from utils.logger import get_logger
from utils.image_utils import decode_base64_image as _decode_image
from config import settings

log = get_logger("omr.scan")
router = APIRouter(prefix="/api/v1/scan", tags=["Tarama"])
limiter = Limiter(key_func=get_remote_address)

GEMINI_API_KEY = os.getenv("GEMINI_API_KEY", "") or settings.gemini_api_key


@router.post("/single", response_model=ScanResponse, summary="Tek kağıt oku")
@limiter.limit("30/minute")
async def scan_single(
    request: Request,
    req: ScanSingleRequest,
    token_data: dict = Depends(verify_firebase_token),
):
    """
    Tek OMR kağıdını okur. Başarılı okumada 1 kredi düşer.
    ArUco tespit başarısız → kredi düşmez, hata döner.
    """
    uid = token_data["uid"]
    goruntu_bytes = _decode_image(req.goruntu_base64)

    # OMR pipeline
    sonuc = kagit_oku(
        goruntu_bytes=goruntu_bytes,
        cevap_anahtari=req.cevap_anahtari,
        api_key=req.gemini_api_key or GEMINI_API_KEY,
        soru_sayisi=req.soru_sayisi,
    )

    # ArUco tespit başarısızdı → kredi düşme
    if sonuc.get("hata") and "ArUco" in str(sonuc.get("hata", "")):
        return ScanResponse(**sonuc)

    # Gemini/başka hata → yine kredi düşme, sonucu döndür
    if sonuc.get("hata"):
        return ScanResponse(**sonuc)

    # Başarılı → 1 kredi düş
    kredi_yeterli = fb.kredi_dус(uid, miktar=1, aciklama=f"Sınav {req.sinav_id} tarama")
    if not kredi_yeterli:
        log.warning("Yetersiz kredi", extra={"uid": uid, "sinav_id": req.sinav_id})
        raise HTTPException(status_code=402, detail="Yetersiz kredi. Lütfen kredi satın alın.")

    log.info("Tarama başarılı — kredi düşüldü", extra={
        "uid": uid, "sinav_id": req.sinav_id,
        "ogrenci_no": sonuc.get("ogrenci_no"), "puan": sonuc.get("puan"),
    })
    # Firestore'a kaydet
    fb.sonuc_kaydet({
        "sinav_id": req.sinav_id,
        "sablon_id": req.sablon_id,
        "ogretmen_id": uid,
        "ogrenci_no": sonuc["ogrenci_no"],
        "ad_soyad": sonuc["ad_soyad"],
        "bolum": sonuc.get("bolum", "?"),
        "ders": sonuc.get("ders", "?"),
        "cevaplar": sonuc["cevaplar"],
        "dogru": sonuc["dogru"],
        "yanlis": sonuc["yanlis"],
        "bos": sonuc["bos"],
        "puan": sonuc["puan"],
        "guvenskor": sonuc["guvenskor"],
    })

    return ScanResponse(**sonuc)


@router.post("/batch", response_model=ScanBatchResponse, summary="Toplu kağıt oku (max 30)")
@limiter.limit("10/minute")
async def scan_batch(
    request: Request,
    req: ScanBatchRequest,
    token_data: dict = Depends(verify_firebase_token),
):
    """
    Birden fazla kağıdı toplu okur.
    Önce kaç kredi gerektiğini hesaplar, yeterliyse devam eder.
    Her başarılı okuma için 1 kredi düşer (bireysel transaction).
    """
    uid = token_data["uid"]

    if len(req.goruntuler) > 30:
        raise HTTPException(status_code=400, detail="Tek istekte maksimum 30 görüntü")

    # Kredi kontrolü
    mevcut_kredi = fb.kredi_oku(uid)
    if mevcut_kredi < 1:
        raise HTTPException(status_code=402, detail="Yetersiz kredi")

    sonuclar: list[BatchSonuc] = []
    kontrol_listesi: list[KontrolGerekli] = []
    basarili = 0
    hatali = 0

    # Paralel işleme
    def _isle(indeks: int, b64: str) -> BatchSonuc:
        try:
            goruntu_bytes = _decode_image(b64)
            sonuc = kagit_oku(
                goruntu_bytes=goruntu_bytes,
                cevap_anahtari=req.cevap_anahtari,
                api_key=GEMINI_API_KEY,
                soru_sayisi=req.soru_sayisi,
            )
            aruco_hatasi = sonuc.get("hata") and "ArUco" in str(sonuc.get("hata", ""))

            if not aruco_hatasi and not sonuc.get("hata"):
                # Kredi düş (bireysel, başarısız olursa okuma yine döner ama kayıt yapılmaz)
                kredi_duste = fb.kredi_dус(uid, miktar=1, aciklama=f"Batch {req.sinav_id}")
                if kredi_duste:
                    fb.sonuc_kaydet({
                        "sinav_id": req.sinav_id,
                        "sablon_id": req.sablon_id,
                        "ogretmen_id": uid,
                        "ogrenci_no": sonuc["ogrenci_no"],
                        "ad_soyad": sonuc["ad_soyad"],
                        "bolum": sonuc.get("bolum", "?"),
                        "ders": sonuc.get("ders", "?"),
                        "cevaplar": sonuc["cevaplar"],
                        "dogru": sonuc["dogru"],
                        "yanlis": sonuc["yanlis"],
                        "bos": sonuc["bos"],
                        "puan": sonuc["puan"],
                        "guvenskor": sonuc["guvenskor"],
                    })

            return BatchSonuc(indeks=indeks, sonuc=ScanResponse(**sonuc))
        except Exception as exc:
            return BatchSonuc(indeks=indeks, hata_mesaji=str(exc))

    with ThreadPoolExecutor(max_workers=settings.thread_workers) as executor:
        futures = {executor.submit(_isle, i, b64): i for i, b64 in enumerate(req.goruntuler)}
        for future in as_completed(futures):
            bs = future.result()
            sonuclar.append(bs)
            if bs.hata_mesaji:
                hatali += 1
                kontrol_listesi.append(KontrolGerekli(
                    indeks=bs.indeks,
                    sebep=bs.hata_mesaji,
                    guvenskor=0.0,
                ))
            else:
                basarili += 1
                sonuc = bs.sonuc
                if sonuc and (sonuc.hata or sonuc.guvenskor < settings.guven_esigi):
                    sebep = sonuc.hata or f"Düşük güven skoru ({sonuc.guvenskor:.2f})"
                    kontrol_listesi.append(KontrolGerekli(
                        indeks=bs.indeks,
                        sebep=sebep,
                        guvenskor=sonuc.guvenskor,
                    ))

    sonuclar.sort(key=lambda x: x.indeks)
    kontrol_listesi.sort(key=lambda x: x.indeks)

    return ScanBatchResponse(
        sonuclar=sonuclar,
        toplam=len(req.goruntuler),
        basarili=basarili,
        hatali=hatali,
        kontrol_gerekli=kontrol_listesi,
    )


# ─────────────────────────── Excel Sınav ────────────────────────────

def _eslesme_durumu(sonuc: dict, og_dict: dict[str, str]) -> str:
    """Web versiyonuyla aynı eşleşme mantığı."""
    if not og_dict:
        return "Liste seçilmedi"
    ogrenci_no = str(sonuc.get("ogrenci_no", "")).strip()
    ad_soyad = str(sonuc.get("ad_soyad", "")).lower()
    no_e = ogrenci_no in og_dict
    liste_ad = og_dict.get(ogrenci_no, "").lower()
    ad_e = any(p in liste_ad for p in ad_soyad.split() if len(p) > 2)
    if no_e and ad_e:
        return "Eşleşme var"
    elif no_e:
        return "No eşleşti, ad farklı"
    elif ad_e:
        return "Ad eşleşti, no farklı"
    else:
        return "Eşleşme yok"


def _excel_ozet(sonuclar: list[dict]) -> str:
    """Özet Excel'i oluştur, base64 döndür."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ozet"
    mavi = PatternFill("solid", fgColor="1a56db")
    beyaz = Font(color="FFFFFF", bold=True)
    kenar = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    basliklar = ["Sayfa", "Ad Soyad", "Ogrenci No", "Durum",
                 "Dogru", "Yanlis", "Bos", "Puan"]
    for j, b in enumerate(basliklar, 1):
        h = ws.cell(row=1, column=j, value=b)
        h.fill = mavi; h.font = beyaz
        h.alignment = Alignment(horizontal="center"); h.border = kenar
    for i, s in enumerate(sonuclar, 2):
        vals = [s.get("sayfa"), s.get("ad_soyad"), s.get("ogrenci_no"),
                s.get("durum"), s.get("dogru"), s.get("yanlis"),
                s.get("bos"), s.get("puan")]
        for j, d in enumerate(vals, 1):
            hc = ws.cell(row=i, column=j, value=d)
            hc.border = kenar
            hc.alignment = Alignment(horizontal="center")
            durum = s.get("durum", "")
            if "Eşleşme var" in durum:
                hc.fill = PatternFill("solid", fgColor="d1fae5")
            elif "farklı" in durum:
                hc.fill = PatternFill("solid", fgColor="fef3c7")
            elif "yok" in durum or "Hata" in durum:
                hc.fill = PatternFill("solid", fgColor="fee2e2")
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18
    with io.BytesIO() as buf:
        wb.save(buf)
        return base64.b64encode(buf.getvalue()).decode()


def _excel_detay(sonuclar: list[dict], cevap_anahtari: dict, soru_sayisi: int) -> str:
    """Detay Excel'i oluştur (her sorunun cevabı), base64 döndür."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Detay"
    mavi = PatternFill("solid", fgColor="1a56db")
    beyaz = Font(color="FFFFFF", bold=True)
    kenar = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    basliklar = ["Ad Soyad", "Ogrenci No", "Puan"] + [f"S{i}" for i in range(1, soru_sayisi + 1)]
    for j, b in enumerate(basliklar, 1):
        h = ws.cell(row=1, column=j, value=b)
        h.fill = mavi; h.font = beyaz
        h.alignment = Alignment(horizontal="center"); h.border = kenar
    # Cevap anahtarı satırı
    ws.cell(row=2, column=1, value="CEVAP ANAHTARI").font = Font(bold=True)
    ws.cell(row=2, column=2, value="-")
    ws.cell(row=2, column=3, value="-")
    for s in range(1, soru_sayisi + 1):
        hc = ws.cell(row=2, column=s + 3, value=cevap_anahtari.get(str(s), "?"))
        hc.fill = PatternFill("solid", fgColor="dbeafe")
        hc.font = Font(bold=True)
        hc.alignment = Alignment(horizontal="center"); hc.border = kenar
    for i, s in enumerate(sonuclar, 3):
        ws.cell(row=i, column=1, value=s.get("ad_soyad", "")).border = kenar
        ws.cell(row=i, column=2, value=s.get("ogrenci_no", "")).border = kenar
        ws.cell(row=i, column=3, value=s.get("puan", 0)).border = kenar
        for soru in range(1, soru_sayisi + 1):
            c = s.get("cevaplar", {}).get(str(soru), "?")
            a = cevap_anahtari.get(str(soru), "?")
            hc = ws.cell(row=i, column=soru + 3, value=c)
            hc.alignment = Alignment(horizontal="center"); hc.border = kenar
            if c == a:
                hc.fill = PatternFill("solid", fgColor="d1fae5")
            elif c == "BOS":
                hc.fill = PatternFill("solid", fgColor="f3f4f6")
            else:
                hc.fill = PatternFill("solid", fgColor="fee2e2")
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 6
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 14
    with io.BytesIO() as buf:
        wb.save(buf)
        return base64.b64encode(buf.getvalue()).decode()


@router.post("/excel-sinav", response_model=ExcelSinavResponse,
             summary="Excel listesiyle toplu tarama")
async def scan_excel_sinav(
    req: ExcelSinavRequest,
    token_data: dict = Depends(verify_firebase_token),
):
    """
    Fotoğraf listesi + öğrenci Excel'i alır, her kağıdı okur,
    öğrenci listesiyle eşleştirir, Özet ve Detay Excel döndürür.
    """
    api_key = req.gemini_api_key or GEMINI_API_KEY

    # Öğrenci listesini yükle (A=No, B=Ad Soyad, başlık yok)
    og_dict: dict[str, str] = {}
    if req.ogrenci_listesi_b64:
        xl_bytes = base64.b64decode(req.ogrenci_listesi_b64)
        xl_wb = openpyxl.load_workbook(io.BytesIO(xl_bytes))
        xl_ws = xl_wb.active
        for row in xl_ws.iter_rows(min_row=1, values_only=True):
            if row[0] is not None and len(row) > 1 and row[1] is not None:
                og_dict[str(row[0]).strip()] = str(row[1]).strip()

    basarili = 0
    hatali = 0
    sonuclar: list[dict] = []

    def _isle(indeks: int, b64: str) -> dict:
        try:
            goruntu_bytes = _decode_image(b64)
            s = kagit_oku(
                goruntu_bytes=goruntu_bytes,
                cevap_anahtari=req.cevap_anahtari,
                api_key=api_key,
                soru_sayisi=req.soru_sayisi,
            )
            s["sayfa"] = indeks + 1
            s["durum"] = "Hata" if s.get("hata") else _eslesme_durumu(s, og_dict)
            return s
        except Exception as exc:
            return {
                "sayfa": indeks + 1, "hata": str(exc), "durum": "Hata",
                "ad_soyad": "?", "ogrenci_no": "?", "dogru": 0,
                "yanlis": 0, "bos": req.soru_sayisi, "puan": 0.0,
                "cevaplar": {}, "guvenskor": 0.0,
            }

    with ThreadPoolExecutor(max_workers=settings.thread_workers) as executor:
        futures = {executor.submit(_isle, i, b64): i
                   for i, b64 in enumerate(req.goruntuler)}
        for future in as_completed(futures):
            s = future.result()
            sonuclar.append(s)
            if s.get("hata"):
                hatali += 1
            else:
                basarili += 1

    sonuclar.sort(key=lambda x: x.get("sayfa", 0))

    return ExcelSinavResponse(
        ozet_excel_b64=_excel_ozet(sonuclar),
        detay_excel_b64=_excel_detay(sonuclar, req.cevap_anahtari, req.soru_sayisi),
        toplam=len(req.goruntuler),
        basarili=basarili,
        hatali=hatali,
    )
