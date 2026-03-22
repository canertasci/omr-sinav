"""
Ortak Excel formatlama yardımcıları.
scan.py ve results.py'deki tekrarlanan Excel kodunu merkezileştirir.
"""
from __future__ import annotations

import io
import base64

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ─── Renk Şemaları ────────────────────────────────────────────────────────────

RENKLER = {
    "mavi_baslik": "1a56db",
    "yesil": "d1fae5",
    "sari": "fef3c7",
    "kirmizi": "fee2e2",
    "acik_mavi": "dbeafe",
    "gri": "f3f4f6",
    # results.py uyumlu
    "yesil_r": "C6EFCE",
    "kirmizi_r": "FFC7CE",
    "gri_r": "D9D9D9",
    "mavi_r": "BDD7EE",
    "sari_r": "FFEB9C",
    "turuncu": "F4B942",
}

# ─── Stil Yardımcıları ─────────────────────────────────────────────────────────

def fill(renk_kodu: str) -> PatternFill:
    return PatternFill("solid", fgColor=renk_kodu)


def ince_kenar() -> Border:
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)


def baslik_stili(ws, satir: int, sutunlar: list[str], renk: str = "mavi_baslik") -> None:
    """Başlık satırını formatla: mavi arka plan + beyaz kalın yazı."""
    for j, baslik_metni in enumerate(sutunlar, 1):
        hucre = ws.cell(row=satir, column=j, value=baslik_metni)
        hucre.fill = fill(RENKLER[renk])
        hucre.font = Font(color="FFFFFF", bold=True)
        hucre.alignment = Alignment(horizontal="center")
        hucre.border = ince_kenar()


def sutun_genislikleri(ws, genislikler: list[int]) -> None:
    """Sütun genişliklerini ayarla."""
    for i, genislik in enumerate(genislikler, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = genislik


# ─── Workbook → Base64 ────────────────────────────────────────────────────────

def workbook_to_b64(wb: openpyxl.Workbook) -> str:
    """Workbook'u bellekte kaydedip base64 string döner."""
    buf = io.BytesIO()
    try:
        wb.save(buf)
        return base64.b64encode(buf.getvalue()).decode()
    finally:
        buf.close()


def workbook_to_bytes(wb: openpyxl.Workbook) -> bytes:
    """Workbook'u bellekte kaydedip bytes döner (StreamingResponse için)."""
    buf = io.BytesIO()
    try:
        wb.save(buf)
        buf.seek(0)
        return buf.read()
    finally:
        buf.close()
