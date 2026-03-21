import streamlit as st
import sqlite3, bcrypt, json, os, base64, time, tempfile
from datetime import datetime
import cv2, numpy as np, requests
import pandas as pd
from pdf2image import convert_from_path
from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO

# ─── AYARLAR ────────────────────────────────────────────────
TARAMA_DPI     = 300
MAX_GORUNTU_PX = 1200
DB_YOLU        = os.path.join(os.getcwd(), "omr.db")
POPPLER_PATH   = r"C:\PYTHON GENEL\OMR\poppler\poppler-24.08.0\Library\bin" if os.name == "nt" else None

ARUCO_DICT   = cv2.aruco.getPredefinedDictionary(cv2.aruco.DICT_4X4_50)
ARUCO_PARAMS = cv2.aruco.DetectorParameters()
ARUCO_DETEK  = cv2.aruco.ArucoDetector(ARUCO_DICT, ARUCO_PARAMS)

# ─── API KEY ─────────────────────────────────────────────────
def _get_gemini_key():
    """Gemini API key'i st.secrets veya env değişkeninden oku."""
    try:
        key = st.secrets.get("GEMINI_API_KEY", "")
        if key:
            return key
    except Exception:
        pass
    return os.getenv("GEMINI_API_KEY", "")

# ─── VERİTABANI ─────────────────────────────────────────────
def db_bag():
    return sqlite3.connect(DB_YOLU)

def db_olustur():
    con = db_bag()
    con.executescript(
        "CREATE TABLE IF NOT EXISTS kullanicilar ("
        "id INTEGER PRIMARY KEY AUTOINCREMENT,"
        "kullanici_adi TEXT UNIQUE NOT NULL,"
        "sifre_hash TEXT NOT NULL,"
        "tam_ad TEXT);"
        "CREATE TABLE IF NOT EXISTS sablonlar ("
        "id INTEGER PRIMARY KEY AUTOINCREMENT,"
        "kullanici_id INTEGER,"
        "ad TEXT NOT NULL,"
        "soru_sayisi INTEGER DEFAULT 20,"
        "tarih TIMESTAMP DEFAULT CURRENT_TIMESTAMP);"
        "CREATE TABLE IF NOT EXISTS cevap_anahtarlari ("
        "id INTEGER PRIMARY KEY AUTOINCREMENT,"
        "kullanici_id INTEGER,"
        "sablon_id INTEGER,"
        "ad TEXT NOT NULL,"
        "cevaplar TEXT NOT NULL,"
        "tarih TIMESTAMP DEFAULT CURRENT_TIMESTAMP);"
        "CREATE TABLE IF NOT EXISTS ogrenci_listeleri ("
        "id INTEGER PRIMARY KEY AUTOINCREMENT,"
        "kullanici_id INTEGER,"
        "ad TEXT NOT NULL,"
        "ogrenciler TEXT NOT NULL,"
        "tarih TIMESTAMP DEFAULT CURRENT_TIMESTAMP);"
        "CREATE TABLE IF NOT EXISTS taramalar ("
        "id INTEGER PRIMARY KEY AUTOINCREMENT,"
        "kullanici_id INTEGER,"
        "anahtar_id INTEGER,"
        "anahtar_adi TEXT,"
        "sablon_adi TEXT,"
        "soru_sayisi INTEGER,"
        "cevap_anahtari TEXT,"
        "toplam_kagit INTEGER,"
        "basarili INTEGER,"
        "tarih TIMESTAMP DEFAULT CURRENT_TIMESTAMP);"
        "CREATE TABLE IF NOT EXISTS tarama_sonuclari ("
        "id INTEGER PRIMARY KEY AUTOINCREMENT,"
        "tarama_id INTEGER,"
        "sayfa INTEGER,"
        "ad_soyad TEXT,"
        "ogrenci_no TEXT,"
        "cevaplar TEXT,"
        "dogru INTEGER,"
        "yanlis INTEGER,"
        "bos INTEGER,"
        "puan REAL,"
        "durum TEXT,"
        "hata TEXT);"
    )
    sifre = bcrypt.hashpw(b"admin123", bcrypt.gensalt()).decode()
    con.execute("INSERT OR IGNORE INTO kullanicilar (kullanici_adi,sifre_hash,tam_ad) VALUES (?,?,?)",
                ("admin", sifre, "Yönetici"))
    con.commit()
    con.close()

# ─── SİLME ONAY YARDIMCISI ──────────────────────────────────
def _sil_butonu(btn_key, label="Sil"):
    """Silme butonunu göster, tıklanınca session_state'e işaretle."""
    if st.button(label, key=f"sil_btn_{btn_key}"):
        st.session_state[f"sil_bekle_{btn_key}"] = True
        st.rerun()

def _sil_onay_goster(btn_key, isim="bu kayıt"):
    """
    Onay dialogunu göster.
    Kullanıcı 'Evet, Sil' tıklarsa True döner.
    İptal veya henüz onay gösterilmiyorsa False döner.
    """
    if not st.session_state.get(f"sil_bekle_{btn_key}", False):
        return False
    st.warning(f"⚠️ **{isim}** silinecek. Bu işlem geri alınamaz!")
    c1, c2 = st.columns(2)
    if c1.button("İptal", key=f"iptal_{btn_key}"):
        st.session_state[f"sil_bekle_{btn_key}"] = False
        st.rerun()
    if c2.button("✓ Evet, Sil", key=f"onayla_{btn_key}"):
        st.session_state[f"sil_bekle_{btn_key}"] = False
        return True
    return False

# ─── GİRİŞ ──────────────────────────────────────────────────
def giris_kontrol(adi, sifre):
    con = db_bag()
    satir = con.execute("SELECT id,sifre_hash,tam_ad FROM kullanicilar WHERE kullanici_adi=?", (adi,)).fetchone()
    con.close()
    if satir and bcrypt.checkpw(sifre.encode(), satir[1].encode()):
        return {"id": satir[0], "kullanici_adi": adi, "tam_ad": satir[2]}
    return None

def giris_sayfasi():
    st.markdown("""
    <style>
    .ana{text-align:center;color:#1a237e;font-size:2.8rem;font-weight:900;margin-top:3rem;letter-spacing:-1px;}
    .alt{text-align:center;color:#5c6bc0;font-size:1rem;margin-top:4px;}
    .uni{text-align:center;color:#9e9e9e;font-size:0.85rem;margin-bottom:2rem;}
    </style>
    <div class="ana">🎓 ÖğretmenAI</div>
    <div class="alt">Sınav Değerlendirme Sistemi</div>
    <div class="uni">Yalova Üniversitesi</div>
    """, unsafe_allow_html=True)
    _, orta, _ = st.columns([1,2,1])
    with orta:
        with st.container(border=True):
            st.subheader("Giriş Yap")
            adi   = st.text_input("Kullanıcı Adı")
            sifre = st.text_input("Şifre", type="password")
            if st.button("Giriş Yap", use_container_width=True, type="primary"):
                k = giris_kontrol(adi, sifre)
                if k:
                    st.session_state.kullanici = k
                    st.session_state.giris_saati = datetime.now().strftime("%H:%M")
                    st.rerun()
                else:
                    st.error("Kullanıcı adı veya şifre hatalı!")
            st.caption("Varsayılan: admin / admin123")

# ─── OMR YARDIMCI FONKSİYONLAR ──────────────────────────────
def pil_to_cv(img):
    return cv2.cvtColor(np.array(img), cv2.COLOR_RGB2BGR)

def b64_yap(cv_img):
    h, w = cv_img.shape[:2]
    if max(h,w) > MAX_GORUNTU_PX:
        oran = MAX_GORUNTU_PX / max(h,w)
        cv_img = cv2.resize(cv_img, (int(w*oran), int(h*oran)))
    _, buf = cv2.imencode(".jpg", cv_img, [cv2.IMWRITE_JPEG_QUALITY, 90])
    return base64.b64encode(buf).decode()

def aruco_tespit(cv_img):
    gri = cv2.cvtColor(cv_img, cv2.COLOR_BGR2GRAY)
    corners, ids, _ = ARUCO_DETEK.detectMarkers(gri)
    if ids is None or len(ids) < 4:
        return None
    m = {}
    for i, mid in enumerate(ids.flatten()):
        if mid in [0,1,2,3]:
            m[int(mid)] = corners[i][0]
    return m if len(m) == 4 else None

def bolgeleri_ayir(cv_img, m):
    dk = {0:m[0][0], 1:m[1][1], 2:m[2][3], 3:m[3][2]}
    xl = int(dk[0][0]); xr = int(dk[1][0])
    yt = int(dk[0][1]); yb = int(dk[2][1])
    xm = int((xl+xr)/2); ym = int((yt+yb)/2)
    return {
        0: cv_img[yt:ym, xl:xm],
        1: cv_img[yt:ym, xm:xr],
        2: cv_img[ym:yb, xl:xm],
        3: cv_img[ym:yb, xm:xr],
    }

def gemini_cagir_web(cv_img, prompt, api_key):
    url  = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent?key={api_key}"
    body = {
        "contents": [{"parts": [
            {"text": prompt},
            {"inline_data": {"mime_type": "image/jpeg", "data": b64_yap(cv_img)}}
        ]}],
        "generationConfig": {"temperature": 0, "thinkingConfig": {"thinkingBudget": 1024}}
    }
    for d in range(3):
        try:
            r = requests.post(url, json=body, timeout=30)
            if r.status_code == 429:
                time.sleep(10); continue
            r.raise_for_status()
            metin = r.json()["candidates"][0]["content"]["parts"][0]["text"].strip()
            if "```" in metin:
                for p in metin.split("```"):
                    p = p.strip()
                    if p.startswith("json"): p = p[4:].strip()
                    if p.startswith("{"): metin = p.strip(); break
            if "{" in metin and "}" in metin:
                metin = metin[metin.index("{"):metin.rindex("}")+1]
            return json.loads(metin)
        except Exception as e:
            if d < 2: time.sleep(2)
            else: return {"hata": str(e)}

PROMPT_NO = (
    "Bu goruntude bir ogrenci numarasi balon tablosu var.\n"
    "9 SUTUN (soldan saga 1-9. hane), 10 SATIR (0-9 rakamlari).\n"
    "Her sutunda YALNIZCA BIR balon dolu olmalidir.\n"
    "Her sutunu BAGIMSIZ degerlendirip EN KOYU balonu bul.\n"
    "Bos balon = sadece ince cember, ici beyaz.\n"
    "Yanlis okuma yapma, dikkatli incele.\n"
    'SADECE JSON don: {"no": "123456789"}'
)

def cevap_prompt(bas, bit):
    return (
        f"Bu goruntude {bas}-{bit} sorularinin cevap balonlari var.\n"
        "Her satir bir soru, satirda A B C D E siklari var.\n"
        "Her satiri BAGIMSIZ degerlendirip EN KOYU balonu bul.\n"
        "Birden fazla esit koyulukta balon varsa hepsini yaz (ornegin \"A/C\").\n"
        "Hic dolu balon yoksa \"BOS\" yaz.\n"
        f"SADECE JSON don: {{\"{bas}\": \"A\", ..., \"{bit}\": \"C\"}}"
    )

def bilgi_prompt():
    return (
        "Bu goruntude ogrenci bilgi formu var.\n"
        "Ad Soyad, Bolum/Program ve Ders Adi alanlarini oku. El yazisi olabilir.\n"
        "SADECE JSON don: {\"ad_soyad\": \"Ad Soyad\", \"bolum\": \"Bolum\", \"ders\": \"Ders\"}"
    )

def kagit_oku_web(pil_img, cevap_anahtari, api_key, soru_sayisi=20):
    cv_img    = pil_to_cv(pil_img)
    markerlar = aruco_tespit(cv_img)
    if markerlar is None:
        return None, "ArUco marker bulunamadı"
    bolgeler = bolgeleri_ayir(cv_img, markerlar)
    bilgi    = gemini_cagir_web(bolgeler[0], bilgi_prompt(), api_key)
    ad_soyad = bilgi.get("ad_soyad","?") if isinstance(bilgi,dict) and "hata" not in bilgi else "?"
    h, w  = bolgeler[1].shape[:2]
    buyuk = cv2.resize(bolgeler[1], (int(w*1.5), int(h*1.5)), interpolation=cv2.INTER_CUBIC)
    no_s  = gemini_cagir_web(buyuk, PROMPT_NO, api_key)
    no    = str(no_s.get("no","?????????")) if isinstance(no_s,dict) and "hata" not in no_s else "?????????"
    no    = no[:9].ljust(9,"?")
    cevaplar = {}
    orta = soru_sayisi // 2
    c1   = gemini_cagir_web(bolgeler[2], cevap_prompt(1, orta), api_key)
    c2   = gemini_cagir_web(bolgeler[3], cevap_prompt(orta+1, soru_sayisi), api_key)
    if isinstance(c1,dict) and "hata" not in c1:
        cevaplar.update({int(k):str(v).upper() for k,v in c1.items() if str(k).isdigit()})
    if isinstance(c2,dict) and "hata" not in c2:
        cevaplar.update({int(k):str(v).upper() for k,v in c2.items() if str(k).isdigit()})
    dogru  = sum(1 for s in range(1,soru_sayisi+1) if cevaplar.get(s)==cevap_anahtari.get(s))
    yanlis = sum(1 for s in range(1,soru_sayisi+1)
                 if cevaplar.get(s) not in (cevap_anahtari.get(s),"BOS","HATA")
                 and "/" not in str(cevaplar.get(s,"")))
    bos  = sum(1 for s in range(1,soru_sayisi+1) if cevaplar.get(s)=="BOS")
    puan = round(dogru * (100/soru_sayisi), 2)
    return {"ad_soyad":ad_soyad,"ogrenci_no":no,"cevaplar":cevaplar,
            "dogru":dogru,"yanlis":yanlis,"bos":bos,"puan":puan}, None

# ─── EXCEL ──────────────────────────────────────────────────
def excel_ozet(sonuclar):
    wb = Workbook(); ws = wb.active; ws.title = "Ozet"
    mavi  = PatternFill("solid", fgColor="1a56db")
    beyaz = Font(color="FFFFFF", bold=True)
    kenar = Border(left=Side(style="thin"),right=Side(style="thin"),
                   top=Side(style="thin"),bottom=Side(style="thin"))
    for j,b in enumerate(["Sayfa","Ad Soyad","Öğrenci No","Durum","Doğru","Yanlış","Boş","Puan"],1):
        h = ws.cell(row=1,column=j,value=b)
        h.fill=mavi; h.font=beyaz; h.alignment=Alignment(horizontal="center"); h.border=kenar
    for i,s in enumerate(sonuclar,2):
        for j,d in enumerate([s.get("sayfa"),s.get("ad_soyad"),s.get("ogrenci_no"),
                               s.get("durum"),s.get("dogru"),s.get("yanlis"),
                               s.get("bos"),s.get("puan")],1):
            hc = ws.cell(row=i,column=j,value=d)
            hc.border=kenar; hc.alignment=Alignment(horizontal="center")
            durum = s.get("durum","")
            if "Eşleşme var" in durum: hc.fill=PatternFill("solid",fgColor="d1fae5")
            elif "farklı" in durum:    hc.fill=PatternFill("solid",fgColor="fef3c7")
            elif "yok" in durum:       hc.fill=PatternFill("solid",fgColor="fee2e2")
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width=18
    buf = BytesIO(); wb.save(buf); return buf.getvalue()

def excel_detay(sonuclar, cevap_anahtari, soru_sayisi=20):
    wb = Workbook(); ws = wb.active; ws.title = "Detay"
    mavi  = PatternFill("solid",fgColor="1a56db")
    beyaz = Font(color="FFFFFF",bold=True)
    kenar = Border(left=Side(style="thin"),right=Side(style="thin"),
                   top=Side(style="thin"),bottom=Side(style="thin"))
    basliklar = ["Ad Soyad","Öğrenci No"] + [f"S{i}" for i in range(1,soru_sayisi+1)]
    for j,b in enumerate(basliklar,1):
        h = ws.cell(row=1,column=j,value=b)
        h.fill=mavi; h.font=beyaz; h.alignment=Alignment(horizontal="center"); h.border=kenar
    ws.cell(row=2,column=1,value="CEVAP ANAHTARI").font=Font(bold=True)
    ws.cell(row=2,column=2,value="-")
    for s in range(1,soru_sayisi+1):
        hc = ws.cell(row=2,column=s+2,value=cevap_anahtari.get(s,"?"))
        hc.fill=PatternFill("solid",fgColor="dbeafe")
        hc.font=Font(bold=True); hc.alignment=Alignment(horizontal="center"); hc.border=kenar
    for i,s in enumerate(sonuclar,3):
        ws.cell(row=i,column=1,value=s.get("ad_soyad","")).border=kenar
        ws.cell(row=i,column=2,value=s.get("ogrenci_no","")).border=kenar
        for soru in range(1,soru_sayisi+1):
            c  = s.get("cevaplar",{}).get(soru,"?")
            a  = cevap_anahtari.get(soru,"?")
            hc = ws.cell(row=i,column=soru+2,value=c)
            hc.alignment=Alignment(horizontal="center"); hc.border=kenar
            if c==a:       hc.fill=PatternFill("solid",fgColor="d1fae5")
            elif c=="BOS": hc.fill=PatternFill("solid",fgColor="f3f4f6")
            else:          hc.fill=PatternFill("solid",fgColor="fee2e2")
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width=12
    buf = BytesIO(); wb.save(buf); return buf.getvalue()

# ─── CSS ────────────────────────────────────────────────────
def _css_uygula():
    st.markdown("""
    <style>
    /* ─── Sidebar: koyu mavi ─── */
    section[data-testid="stSidebar"] > div:first-child {
        background: linear-gradient(180deg, #1a237e 0%, #283593 100%);
    }
    section[data-testid="stSidebar"] .stMarkdown p,
    section[data-testid="stSidebar"] .stCaption p,
    section[data-testid="stSidebar"] p,
    section[data-testid="stSidebar"] span,
    section[data-testid="stSidebar"] small,
    section[data-testid="stSidebar"] label { color: white !important; }
    section[data-testid="stSidebar"] hr { border-color: rgba(255,255,255,0.25) !important; }
    section[data-testid="stSidebar"] .stRadio div[role="radiogroup"] label {
        color: white !important; padding: 4px 8px; border-radius: 6px; transition: background 0.15s;
    }
    section[data-testid="stSidebar"] .stRadio div[role="radiogroup"] label:hover {
        background: rgba(255,255,255,0.15) !important;
    }
    section[data-testid="stSidebar"] .stButton > button {
        background: rgba(255,255,255,0.15) !important;
        border: 1px solid rgba(255,255,255,0.35) !important;
        color: white !important; border-radius: 8px !important;
    }
    section[data-testid="stSidebar"] .stButton > button:hover {
        background: rgba(255,255,255,0.28) !important;
    }
    /* ─── Primary buttons: yeşil ─── */
    .stButton > button[kind="primary"],
    .stButton > button[kind="primaryFormSubmit"] {
        background: #2E7D32 !important;
        border-color: #2E7D32 !important;
        color: white !important; border-radius: 8px !important;
    }
    .stButton > button[kind="primary"]:hover { background: #1B5E20 !important; }
    /* ─── Secondary/sil buttons: kırmızı ─── */
    div[data-testid="stButton"] button[kind="secondary"] {
        border-color: #C62828 !important; color: #C62828 !important; border-radius: 8px !important;
    }
    div[data-testid="stButton"] button[kind="secondary"]:hover {
        background: #FFEBEE !important; color: #B71C1C !important;
    }
    /* ─── Metric kartları ─── */
    [data-testid="metric-container"] {
        background: white; border: 1px solid #e8eaf6; border-radius: 12px;
        padding: 16px 20px; box-shadow: 0 2px 8px rgba(26,35,126,0.09);
    }
    [data-testid="metric-container"] [data-testid="stMetricLabel"] { color: #5c6bc0 !important; }
    [data-testid="metric-container"] [data-testid="stMetricValue"] {
        color: #1a237e !important; font-weight: 700 !important;
    }
    /* ─── Genel ─── */
    .stApp { background: #f5f6fa; }
    h1, h2, h3 { color: #1a237e !important; }
    </style>
    """, unsafe_allow_html=True)

# ─── ANA SAYFA ───────────────────────────────────────────────
@st.cache_data(ttl=60)
def _dashboard_verileri(uid):
    con = db_bag()
    sablon_sayisi = con.execute("SELECT COUNT(*) FROM sablonlar WHERE kullanici_id=?", (uid,)).fetchone()[0]
    tarama_sayisi = con.execute("SELECT COUNT(*) FROM taramalar WHERE kullanici_id=?", (uid,)).fetchone()[0]
    listeler      = con.execute("SELECT ogrenciler FROM ogrenci_listeleri WHERE kullanici_id=?", (uid,)).fetchall()
    son_taramalar = con.execute(
        "SELECT t.anahtar_adi, t.sablon_adi, t.toplam_kagit, t.basarili, t.tarih, "
        "ROUND(COALESCE(AVG(s.puan),0),1) "
        "FROM taramalar t LEFT JOIN tarama_sonuclari s ON t.id=s.tarama_id "
        "WHERE t.kullanici_id=? GROUP BY t.id ORDER BY t.id DESC LIMIT 5", (uid,)
    ).fetchall()
    con.close()
    ogrenci_sayisi = sum(len(json.loads(r[0])) for r in listeler)
    return sablon_sayisi, tarama_sayisi, ogrenci_sayisi, son_taramalar

def sayfa_anasayfa():
    st.header("Ana Sayfa")
    st.divider()
    uid = st.session_state.kullanici["id"]
    try:
        sablon_sayisi, tarama_sayisi, ogrenci_sayisi, son_taramalar = _dashboard_verileri(uid)

        c1, c2, c3 = st.columns(3)
        c1.metric("📐 Şablon Sayısı", sablon_sayisi,
                  help="Oluşturduğunuz sınav şablonlarının sayısı (10/20/30/40/50 soruluk)")
        c2.metric("👥 Toplam Öğrenci", ogrenci_sayisi,
                  help="Tüm öğrenci listelerindeki toplam kayıtlı öğrenci sayısı")
        c3.metric("📊 Toplam Tarama", tarama_sayisi,
                  help="Şimdiye kadar gerçekleştirilen toplam sınav taraması sayısı")

        st.divider()
        st.subheader("Son 5 Tarama")
        if son_taramalar:
            df = pd.DataFrame(son_taramalar,
                              columns=["Cevap Anahtarı","Şablon","Toplam Kağıt","Başarılı","Tarih","Ort. Puan"])
            df["Tarih"] = df["Tarih"].str[:16]
            st.dataframe(df, use_container_width=True, hide_index=True)
        else:
            st.info("Henüz tarama yapılmamış. 'Sınav Oku' sayfasından başlayabilirsiniz.")
    except Exception as e:
        st.error(f"Hata: {e}")

# ─── SAYFALAR ───────────────────────────────────────────────
def sayfa_sablon():
    st.header("Şablon Yönetimi")
    st.divider()
    uid = st.session_state.kullanici["id"]
    try:
        with st.container(border=True):
            st.subheader("Yeni Şablon Ekle")
            adi = st.text_input("Şablon Adı")
            ss  = st.selectbox("Soru Sayısı", [10,20,30,40,50], index=1)
            if st.button("Kaydet", type="primary"):
                if adi:
                    con = db_bag()
                    con.execute("INSERT INTO sablonlar (kullanici_id,ad,soru_sayisi) VALUES (?,?,?)",(uid,adi,ss))
                    con.commit(); con.close()
                    st.success(f"'{adi}' kaydedildi!"); st.rerun()
                else: st.warning("Şablon adı girin!")
        st.subheader("Kayıtlı Şablonlar")
        con = db_bag()
        rows = con.execute("SELECT id,ad,soru_sayisi,tarih FROM sablonlar WHERE kullanici_id=? ORDER BY id DESC",(uid,)).fetchall()
        con.close()
        if rows:
            for r in rows:
                c1,c2,c3,c4 = st.columns([3,2,3,1])
                c1.markdown(f"**{r[1]}**"); c2.write(f"{r[2]} soru"); c3.write(r[3][:16])
                with c4:
                    _sil_butonu(f"sablon_{r[0]}")
                if _sil_onay_goster(f"sablon_{r[0]}", f"'{r[1]}'"):
                    con2=db_bag(); con2.execute("DELETE FROM sablonlar WHERE id=?",(r[0],)); con2.commit(); con2.close(); st.rerun()
        else: st.info("Henüz şablon yok.")
    except Exception as e:
        st.error(f"Hata: {e}")

def sayfa_anahtar():
    st.header("Cevap Anahtarı")
    st.divider()
    uid = st.session_state.kullanici["id"]
    try:
        con = db_bag()
        sablonlar = con.execute("SELECT id,ad,soru_sayisi FROM sablonlar WHERE kullanici_id=?",(uid,)).fetchall()
        con.close()
        if not sablonlar: st.warning("Önce şablon ekleyin!"); return
        sablon = st.selectbox("Şablon Seç", sablonlar, format_func=lambda x: f"{x[1]} ({x[2]} soru)")
        ss     = sablon[2]
        adi    = st.text_input("Cevap Anahtarı Adı (örn: Vize 2025)")

        tab1, tab2 = st.tabs(["✏️ Manuel Giriş", "📂 Dosyadan Yükle"])
        cevaplar = {}

        with tab1:
            st.caption("Her soru için doğru şıkkı seçin.")
            for i in range(0, ss, 5):
                cols = st.columns(5)
                for j,col in enumerate(cols):
                    sno = i+j+1
                    if sno <= ss:
                        with col:
                            cevaplar[sno] = st.selectbox(f"Soru {sno}",["A","B","C","D","E"],key=f"ca{sno}")

        with tab2:
            # Örnek şablon indir
            ornek_df = pd.DataFrame({"Soru No": range(1, ss+1), "Cevap": ["A"]*ss})
            ornek_buf = BytesIO()
            ornek_df.to_excel(ornek_buf, index=False, engine="openpyxl")
            st.download_button("⬇️ Örnek Şablon İndir", ornek_buf.getvalue(),
                               "ornek_cevap_anahtari.xlsx",
                               "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            st.caption("A sütunu = Soru No, B sütunu = Cevap (A/B/C/D/E)")
            yuklenen = st.file_uploader("CSV veya Excel Yükle", type=["csv","xlsx","xls"],
                                        key="anahtar_yukle")
            if yuklenen:
                try:
                    if yuklenen.name.endswith(".csv"):
                        df_yukle = pd.read_csv(yuklenen, header=0)
                    else:
                        df_yukle = pd.read_excel(yuklenen, header=0, engine="openpyxl")
                    df_yukle.columns = [str(c).strip() for c in df_yukle.columns]
                    # İlk sütun soru no, ikinci sütun cevap
                    for _, row in df_yukle.iterrows():
                        try:
                            sno = int(row.iloc[0])
                            cvp = str(row.iloc[1]).strip().upper()
                            if 1 <= sno <= ss and cvp in ["A","B","C","D","E"]:
                                cevaplar[sno] = cvp
                        except: pass
                    st.success(f"{len(cevaplar)} soru yüklendi.")
                    st.dataframe(pd.DataFrame([{"Soru":k,"Cevap":v} for k,v in sorted(cevaplar.items())]),
                                 hide_index=True)
                except Exception as e:
                    st.error(f"Dosya okunamadı: {e}")

        if st.button("Anahtarı Kaydet", type="primary"):
            if adi:
                con = db_bag()
                con.execute("INSERT INTO cevap_anahtarlari (kullanici_id,sablon_id,ad,cevaplar) VALUES (?,?,?,?)",
                            (uid,sablon[0],adi,json.dumps(cevaplar)))
                con.commit(); con.close()
                st.success(f"'{adi}' kaydedildi!")
            else: st.warning("Cevap anahtarı adı girin!")
        st.subheader("Kayıtlı Anahtarlar")
        con = db_bag()
        rows = con.execute("SELECT id,ad,sablon_id,cevaplar,tarih FROM cevap_anahtarlari WHERE kullanici_id=? ORDER BY id DESC",(uid,)).fetchall()
        con.close()
        for r in rows:
            c1,c2,c3,c4 = st.columns([4,2,1,1])
            c1.write(f"**{r[1]}**"); c2.write(r[4][:16])
            duzenle_key = f"duzenle_{r[0]}"
            if c3.button("✏️", key=f"btn_{duzenle_key}", help="Düzenle"):
                st.session_state[duzenle_key] = not st.session_state.get(duzenle_key, False)
                st.rerun()
            with c4:
                _sil_butonu(f"anahtar_{r[0]}")
            if _sil_onay_goster(f"anahtar_{r[0]}", f"'{r[1]}'"):
                con2=db_bag(); con2.execute("DELETE FROM cevap_anahtarlari WHERE id=?",(r[0],)); con2.commit(); con2.close(); st.rerun()

            # Düzenleme paneli
            if st.session_state.get(duzenle_key, False):
                mevcut = {int(k):v for k,v in json.loads(r[3]).items()}
                # Şablonun soru sayısını bul
                con3 = db_bag()
                sablon_row = con3.execute("SELECT soru_sayisi FROM sablonlar WHERE id=?", (r[2],)).fetchone()
                con3.close()
                edit_ss = sablon_row[0] if sablon_row else len(mevcut)
                with st.container(border=True):
                    st.markdown(f"**✏️ Düzenleniyor: {r[1]}**")
                    yeni_cevaplar = {}
                    for i in range(0, edit_ss, 5):
                        cols = st.columns(5)
                        for j, col in enumerate(cols):
                            sno = i+j+1
                            if sno <= edit_ss:
                                with col:
                                    varsayilan = ["A","B","C","D","E"].index(mevcut.get(sno,"A"))
                                    yeni_cevaplar[sno] = st.selectbox(
                                        f"Soru {sno}", ["A","B","C","D","E"],
                                        index=varsayilan, key=f"edit_{r[0]}_{sno}"
                                    )
                    col_kaydet, col_iptal = st.columns(2)
                    if col_kaydet.button("💾 Kaydet", key=f"kaydet_{r[0]}", type="primary"):
                        con4 = db_bag()
                        con4.execute("UPDATE cevap_anahtarlari SET cevaplar=? WHERE id=?",
                                     (json.dumps(yeni_cevaplar), r[0]))
                        con4.commit(); con4.close()
                        st.session_state[duzenle_key] = False
                        st.success("Güncellendi!"); st.rerun()
                    if col_iptal.button("İptal", key=f"iptal_edit_{r[0]}"):
                        st.session_state[duzenle_key] = False
                        st.rerun()
    except Exception as e:
        st.error(f"Hata: {e}")

def sayfa_liste():
    st.header("Öğrenci Listesi")
    st.divider()
    uid = st.session_state.kullanici["id"]
    try:
        with st.container(border=True):
            st.subheader("Excel Yükle")
            st.info("📋 **Excel formatı:** Başlık satırı olmadan, **A sütunu** = Öğrenci No, **B sütunu** = Ad Soyad")
            ornek_og = pd.DataFrame([
                {"Öğrenci No": "22010001", "Ad Soyad": "Ali Yılmaz"},
                {"Öğrenci No": "22010002", "Ad Soyad": "Ayşe Kaya"},
            ])
            ornek_buf = BytesIO()
            ornek_og.to_excel(ornek_buf, index=False, header=False, engine="openpyxl")
            st.download_button("⬇️ Örnek Excel Şablonu İndir", ornek_buf.getvalue(),
                               "ornek_ogrenci_listesi.xlsx",
                               "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               key="ornek_og_indir")
            adi      = st.text_input("Liste Adı")
            yuklenen = st.file_uploader("Excel Dosyası", type=["xlsx","xls"])
            if yuklenen and adi and st.button("Listeyi Kaydet", type="primary"):
                engine = "openpyxl" if yuklenen.name.endswith(".xlsx") else "xlrd"
                df = pd.read_excel(yuklenen, header=None, engine=engine)
                ogrenciler = [{"no":str(r[0]).strip(),"ad":str(r[1]).strip()} for _,r in df.iterrows()]
                con = db_bag()
                con.execute("INSERT INTO ogrenci_listeleri (kullanici_id,ad,ogrenciler) VALUES (?,?,?)",
                            (uid,adi,json.dumps(ogrenciler,ensure_ascii=False)))
                con.commit(); con.close()
                st.success(f"{len(ogrenciler)} öğrenci kaydedildi!"); st.rerun()
        st.subheader("Kayıtlı Listeler")
        con = db_bag()
        rows = con.execute("SELECT id,ad,ogrenciler,tarih FROM ogrenci_listeleri WHERE kullanici_id=? ORDER BY id DESC",(uid,)).fetchall()
        con.close()
        for r in rows:
            og = json.loads(r[2])
            c1,c2,c3,c4,c5 = st.columns([3,2,2,1,1])
            c1.markdown(f"**{r[1]}**"); c2.write(f"{len(og)} öğrenci"); c3.write(r[3][:16])
            goruntu_key = f"goruntu_{r[0]}"
            if c4.button("👁️", key=f"btn_{goruntu_key}", help="Görüntüle"):
                st.session_state[goruntu_key] = not st.session_state.get(goruntu_key, False)
                st.rerun()
            with c5:
                _sil_butonu(f"liste_{r[0]}")
            if _sil_onay_goster(f"liste_{r[0]}", f"'{r[1]}'"):
                con2=db_bag(); con2.execute("DELETE FROM ogrenci_listeleri WHERE id=?",(r[0],)); con2.commit(); con2.close(); st.rerun()

            if st.session_state.get(goruntu_key, False):
                with st.container(border=True):
                    st.markdown(f"**👥 {r[1]}** — {len(og)} öğrenci")
                    df_og = pd.DataFrame(og).rename(columns={"no": "Öğrenci No", "ad": "Ad Soyad"})
                    st.dataframe(df_og, use_container_width=True, hide_index=True)
                    buf = BytesIO()
                    df_og.to_excel(buf, index=False, engine="openpyxl")
                    st.download_button("⬇️ Excel İndir", buf.getvalue(),
                                       f"{r[1]}.xlsx",
                                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                       key=f"indir_{r[0]}")
    except Exception as e:
        st.error(f"Hata: {e}")

def sayfa_sinav():
    st.header("Sınav Oku")
    st.divider()
    uid = st.session_state.kullanici["id"]
    try:
        con = db_bag()
        sablonlar  = con.execute("SELECT id,ad,soru_sayisi FROM sablonlar WHERE kullanici_id=?",(uid,)).fetchall()
        anahtarlar = con.execute("SELECT id,ad,sablon_id,cevaplar FROM cevap_anahtarlari WHERE kullanici_id=?",(uid,)).fetchall()
        listeler   = con.execute("SELECT id,ad,ogrenciler FROM ogrenci_listeleri WHERE kullanici_id=?",(uid,)).fetchall()
        con.close()
        if not sablonlar or not anahtarlar:
            st.warning("Önce şablon ve cevap anahtarı ekleyin!"); return
        c1,c2 = st.columns(2)
        with c1:
            sablon  = st.selectbox("Şablon", sablonlar, format_func=lambda x: f"{x[1]} ({x[2]} soru)")
            anahtar = st.selectbox("Cevap Anahtarı", anahtarlar, format_func=lambda x: x[1])
        with c2:
            liste = st.selectbox("Öğrenci Listesi (opsiyonel)", [None]+list(listeler),
                                 format_func=lambda x: "Seçme" if x is None else x[1])
            # API Key: secrets/env'den al, yoksa kullanıcıdan iste
            _api_key_sabit = _get_gemini_key()
            if _api_key_sabit:
                api_key = _api_key_sabit
                st.info("✅ Gemini API Key yapılandırılmış.")
            else:
                api_key = st.text_input("Gemini API Key", type="password",
                                        placeholder="AIza...")
                with st.expander("ℹ️ Gemini API Key nasıl alınır?"):
                    st.markdown("""
1. [aistudio.google.com](https://aistudio.google.com) adresine git
2. Google hesabınla giriş yap
3. Sol menüden **"Get API Key"** → **"Create API Key"** tıkla
4. Oluşturulan `AIza...` ile başlayan anahtarı kopyala
5. Yukarıdaki alana yapıştır

> 🔒 Key sadece bu oturumda kullanılır, sunucuda saklanmaz.
                    """)
        pdf = st.file_uploader("Sınav PDF", type=["pdf"])
        if pdf and api_key and st.button("Sınavı Oku", type="primary", use_container_width=True):
            anahtardict = {int(k):v for k,v in json.loads(anahtar[3]).items()}
            ss          = sablon[2]
            og_dict     = {}
            if liste:
                og_dict = {o["no"]:o["ad"] for o in json.loads(liste[2])}
            with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as f:
                f.write(pdf.read()); tmp = f.name
            with st.spinner("PDF dönüştürülüyor..."):
                sayfalar = convert_from_path(tmp, dpi=TARAMA_DPI, poppler_path=POPPLER_PATH)
            toplam   = len(sayfalar)
            st.write(f"**{toplam} sayfa bulundu**")
            prog     = st.progress(0)
            durum    = st.empty()
            sonuclar = []
            with st.spinner("Sayfalar Gemini ile okunuyor..."):
                for i,sayfa in enumerate(sayfalar,1):
                    if i > 1: time.sleep(1)
                    durum.write(f"Sayfa {i}/{toplam} okunuyor...")
                    s, hata = kagit_oku_web(sayfa, anahtardict, api_key, ss)
                    if hata:
                        sonuclar.append({"sayfa":i,"hata":hata,"durum":"Hata",
                                          "ad_soyad":"?","ogrenci_no":"?","dogru":0,
                                          "yanlis":0,"bos":ss,"puan":0,"cevaplar":{}})
                    else:
                        d = "Liste seçilmedi"
                        if og_dict:
                            no_e     = s["ogrenci_no"] in og_dict
                            liste_ad = og_dict.get(s["ogrenci_no"],"").lower()
                            ad_e     = any(p in liste_ad for p in s["ad_soyad"].lower().split() if len(p)>2)
                            if no_e and ad_e:  d = "Eşleşme var"
                            elif no_e:         d = "No eşleşti, ad farklı"
                            elif ad_e:         d = "Ad eşleşti, no farklı"
                            else:              d = "Eşleşme yok"
                        s["sayfa"] = i; s["durum"] = d
                        sonuclar.append(s)
                    prog.progress(i/toplam)
            durum.empty()
            basarili = sum(1 for s in sonuclar if not s.get("hata"))
            con = db_bag()
            cur = con.execute(
                "INSERT INTO taramalar (kullanici_id,anahtar_id,anahtar_adi,sablon_adi,"
                "soru_sayisi,cevap_anahtari,toplam_kagit,basarili) VALUES (?,?,?,?,?,?,?,?)",
                (uid, anahtar[0], anahtar[1], sablon[1], ss,
                 json.dumps(anahtardict), toplam, basarili)
            )
            tarama_id = cur.lastrowid
            for s in sonuclar:
                con.execute(
                    "INSERT INTO tarama_sonuclari (tarama_id,sayfa,ad_soyad,ogrenci_no,"
                    "cevaplar,dogru,yanlis,bos,puan,durum,hata) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                    (tarama_id, s.get("sayfa"), s.get("ad_soyad","?"),
                     s.get("ogrenci_no","?"), json.dumps(s.get("cevaplar",{})),
                     s.get("dogru",0), s.get("yanlis",0), s.get("bos",0),
                     s.get("puan",0), s.get("durum",""), s.get("hata"))
                )
            con.commit(); con.close()
            st.session_state.sonuclar    = sonuclar
            st.session_state.anahtardict = anahtardict
            st.session_state.ss          = ss
            st.success(f"{toplam} kağıt okundu ve kaydedildi!"); st.rerun()
        if "sonuclar" in st.session_state:
            sonuclar    = st.session_state.sonuclar
            anahtardict = st.session_state.anahtardict
            ss          = st.session_state.ss
            st.subheader("Sonuçlar")
            df_data = [{"Sayfa":s.get("sayfa"),"Ad Soyad":s.get("ad_soyad"),
                        "No":s.get("ogrenci_no"),"Durum":s.get("durum"),
                        "Doğru":s.get("dogru"),"Yanlış":s.get("yanlis"),
                        "Puan":s.get("puan")} for s in sonuclar]
            st.dataframe(pd.DataFrame(df_data), use_container_width=True)
            c1,c2 = st.columns(2)
            with c1:
                st.download_button("Özet Excel İndir", excel_ozet(sonuclar),
                                   "ozet.xlsx","application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                   use_container_width=True)
            with c2:
                st.download_button("Detay Excel İndir", excel_detay(sonuclar,anahtardict,ss),
                                   "detay.xlsx","application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                   use_container_width=True)
    except Exception as e:
        st.error(f"Hata: {e}")

def sayfa_gecmis():
    st.header("Geçmiş Taramalar")
    st.divider()
    uid = st.session_state.kullanici["id"]
    try:
        con = db_bag()
        taramalar = con.execute(
            "SELECT id,anahtar_adi,sablon_adi,soru_sayisi,toplam_kagit,basarili,tarih "
            "FROM taramalar WHERE kullanici_id=? ORDER BY id DESC", (uid,)
        ).fetchall()
        con.close()

        if not taramalar:
            st.info("Henüz kayıtlı tarama yok. 'Sınav Oku' sayfasından tarama yapabilirsiniz.")
            return

        if "gecmis_secili" not in st.session_state:
            st.session_state.gecmis_secili = None

        # ── Filtreleme ───────────────────────────────────────
        with st.expander("🔍 Filtrele", expanded=False):
            fc1, fc2 = st.columns(2)
            with fc1:
                sablon_adlari = sorted({t[2] for t in taramalar})
                secili_sablon = st.selectbox("Şablon Adı", ["Tümü"] + sablon_adlari)
            with fc2:
                tarih_sec = st.date_input("Tarih Filtresi", value=None, key="gcm_tarih")
                tarih_filtre = tarih_sec.strftime("%Y-%m") if tarih_sec else ""

        if secili_sablon != "Tümü":
            taramalar = [t for t in taramalar if t[2] == secili_sablon]
        if tarih_filtre:
            taramalar = [t for t in taramalar if tarih_filtre in t[6]]

        # ── Toplu Excel Export ───────────────────────────────
        if taramalar:
            if st.button("📥 Tümünü Excel'e Aktar", type="primary"):
                rows = []
                for t in taramalar:
                    rows.append({
                        "Cevap Anahtarı": t[1], "Şablon": t[2],
                        "Soru Sayısı": t[3], "Toplam Kağıt": t[4],
                        "Başarılı": t[5], "Tarih": t[6][:16]
                    })
                df_exp = pd.DataFrame(rows)
                buf = BytesIO()
                df_exp.to_excel(buf, index=False, engine="openpyxl")
                st.download_button(
                    "⬇️ İndir", buf.getvalue(), "gecmis_taramalar.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

        st.subheader(f"Toplam {len(taramalar)} tarama")
        for t in taramalar:
            tid, anahtar_adi, sablon_adi, ss, toplam, basarili, tarih = t
            hata_sayisi = toplam - basarili
            secili = st.session_state.gecmis_secili == tid

            with st.container(border=True):
                c1, c2, c3, c4, c5 = st.columns([3, 2, 2, 1, 1])
                c1.markdown(f"**{anahtar_adi}**  \n{sablon_adi} · {ss} soru")
                c2.markdown(f"**{toplam}** kağıt  \n{basarili} başarılı" +
                            (f", {hata_sayisi} hata" if hata_sayisi else ""))
                c3.markdown(f"{tarih[:16]}")
                if c4.button("Göster" if not secili else "Gizle", key=f"gos{tid}"):
                    st.session_state.gecmis_secili = None if secili else tid
                    st.rerun()
                with c5:
                    _sil_butonu(f"tarama_{tid}")
                if _sil_onay_goster(f"tarama_{tid}", f"'{anahtar_adi}' taraması"):
                    con2 = db_bag()
                    con2.execute("DELETE FROM tarama_sonuclari WHERE tarama_id=?", (tid,))
                    con2.execute("DELETE FROM taramalar WHERE id=?", (tid,))
                    con2.commit(); con2.close()
                    if st.session_state.gecmis_secili == tid:
                        st.session_state.gecmis_secili = None
                    st.rerun()

            if secili:
                con = db_bag()
                row_t = con.execute(
                    "SELECT cevap_anahtari FROM taramalar WHERE id=?", (tid,)
                ).fetchone()
                sonuclar_db = con.execute(
                    "SELECT sayfa,ad_soyad,ogrenci_no,cevaplar,dogru,yanlis,bos,puan,durum,hata "
                    "FROM tarama_sonuclari WHERE tarama_id=? ORDER BY sayfa", (tid,)
                ).fetchall()
                con.close()

                anahtardict = {int(k): v for k, v in json.loads(row_t[0]).items()}
                sonuclar = []
                for r in sonuclar_db:
                    sonuclar.append({
                        "sayfa": r[0], "ad_soyad": r[1], "ogrenci_no": r[2],
                        "cevaplar": {int(k): v for k, v in json.loads(r[3]).items()},
                        "dogru": r[4], "yanlis": r[5], "bos": r[6],
                        "puan": r[7], "durum": r[8], "hata": r[9],
                    })

                df_data = [{"Sayfa": s["sayfa"], "Ad Soyad": s["ad_soyad"],
                            "No": s["ogrenci_no"], "Durum": s["durum"],
                            "Doğru": s["dogru"], "Yanlış": s["yanlis"], "Puan": s["puan"]}
                           for s in sonuclar]
                st.dataframe(pd.DataFrame(df_data), use_container_width=True)

                ec1, ec2 = st.columns(2)
                with ec1:
                    st.download_button(
                        "Özet Excel", excel_ozet(sonuclar), f"ozet_{tid}.xlsx",
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True, key=f"exc_ozet_{tid}"
                    )
                with ec2:
                    st.download_button(
                        "Detay Excel", excel_detay(sonuclar, anahtardict, ss), f"detay_{tid}.xlsx",
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True, key=f"exc_detay_{tid}"
                    )
    except Exception as e:
        st.error(f"Hata: {e}")

def sayfa_kullanici():
    st.header("Kullanıcı Yönetimi")
    st.divider()
    uid  = st.session_state.kullanici["id"]
    kadi = st.session_state.kullanici["kullanici_adi"]
    try:
        with st.container(border=True):
            st.subheader("Yeni Kullanıcı Ekle")
            yadi = st.text_input("Kullanıcı Adı")
            ytam = st.text_input("Ad Soyad")
            ysif = st.text_input("Şifre", type="password")
            if st.button("Kullanıcı Ekle", type="primary"):
                if yadi and ysif:
                    try:
                        sh = bcrypt.hashpw(ysif.encode(),bcrypt.gensalt()).decode()
                        con = db_bag()
                        con.execute("INSERT INTO kullanicilar (kullanici_adi,sifre_hash,tam_ad) VALUES (?,?,?)",(yadi,sh,ytam))
                        con.commit(); con.close()
                        st.success(f"'{yadi}' eklendi!")
                    except: st.error("Bu kullanıcı adı zaten var!")
                else: st.warning("Kullanıcı adı ve şifre gerekli!")
        with st.container(border=True):
            st.subheader("Şifre Değiştir")
            eski = st.text_input("Mevcut Şifre", type="password")
            yeni = st.text_input("Yeni Şifre", type="password")
            if st.button("Şifreyi Değiştir"):
                if giris_kontrol(kadi, eski):
                    sh = bcrypt.hashpw(yeni.encode(),bcrypt.gensalt()).decode()
                    con = db_bag()
                    con.execute("UPDATE kullanicilar SET sifre_hash=? WHERE id=?",(sh,uid))
                    con.commit(); con.close()
                    st.success("Şifre değiştirildi!")
                else: st.error("Mevcut şifre hatalı!")
    except Exception as e:
        st.error(f"Hata: {e}")

# ─── ANA UYGULAMA ────────────────────────────────────────────
st.set_page_config(page_title="ÖğretmenAI | OMR Sistemi", page_icon="🎓",
                   layout="wide", initial_sidebar_state="expanded")
db_olustur()

if "kullanici" not in st.session_state:
    giris_sayfasi()
else:
    _css_uygula()
    k = st.session_state.kullanici
    with st.sidebar:
        st.markdown("""
        <div style="text-align:center;padding:1.2rem 0 0.5rem;">
            <div style="font-size:1.7rem;font-weight:900;letter-spacing:-0.5px;">🎓 ÖğretmenAI</div>
            <div style="font-size:0.72rem;opacity:0.85;margin-top:3px;">Sınav Değerlendirme Sistemi</div>
            <div style="font-size:0.68rem;opacity:0.65;margin-top:2px;">Yalova Üniversitesi</div>
        </div>
        """, unsafe_allow_html=True)
        st.divider()
        tam_ad = k['tam_ad'] or k['kullanici_adi']
        giris_saati = st.session_state.get("giris_saati", "")
        st.markdown(f"**👤 {tam_ad}**")
        if giris_saati:
            st.caption(f"🕐 Giriş: {giris_saati}")
        st.divider()
        sayfa = st.radio("Menü", [
            "🏠 Ana Sayfa",
            "📐 Şablon Yönetimi",
            "🔑 Cevap Anahtarı",
            "👥 Öğrenci Listesi",
            "📋 Sınav Oku",
            "📊 Geçmiş Taramalar",
            "⚙️ Kullanıcı Yönetimi",
        ], label_visibility="collapsed")
        st.divider()
        if st.button("🚪 Çıkış Yap", use_container_width=True):
            del st.session_state.kullanici
            st.rerun()

    if   sayfa == "🏠 Ana Sayfa":            sayfa_anasayfa()
    elif sayfa == "📐 Şablon Yönetimi":      sayfa_sablon()
    elif sayfa == "🔑 Cevap Anahtarı":       sayfa_anahtar()
    elif sayfa == "👥 Öğrenci Listesi":      sayfa_liste()
    elif sayfa == "📋 Sınav Oku":            sayfa_sinav()
    elif sayfa == "📊 Geçmiş Taramalar":     sayfa_gecmis()
    elif sayfa == "⚙️ Kullanıcı Yönetimi":  sayfa_kullanici()
