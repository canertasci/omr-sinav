"""
Excel export yardımcı fonksiyonları — tüm sayfalar buradan import eder.
"""
from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def excel_ozet(sonuclar: list[dict]) -> bytes:
    """Özet (Özet) sayfası: her öğrenci için tek satır."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Ozet"
    mavi  = PatternFill("solid", fgColor="1a56db")
    beyaz = Font(color="FFFFFF", bold=True)
    kenar = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    for j, b in enumerate(
        ["Sayfa", "Ad Soyad", "Öğrenci No", "Durum", "Doğru", "Yanlış", "Boş", "Puan"], 1
    ):
        h = ws.cell(row=1, column=j, value=b)
        h.fill = mavi
        h.font = beyaz
        h.alignment = Alignment(horizontal="center")
        h.border = kenar
    for i, s in enumerate(sonuclar, 2):
        for j, d in enumerate(
            [s.get("sayfa"), s.get("ad_soyad"), s.get("ogrenci_no"),
             s.get("durum"), s.get("dogru"), s.get("yanlis"),
             s.get("bos"), s.get("puan")], 1,
        ):
            hc = ws.cell(row=i, column=j, value=d)
            hc.border = kenar
            hc.alignment = Alignment(horizontal="center")
            durum = s.get("durum", "")
            if "Eşleşme var" in durum:
                hc.fill = PatternFill("solid", fgColor="d1fae5")
            elif "farklı" in durum:
                hc.fill = PatternFill("solid", fgColor="fef3c7")
            elif "yok" in durum:
                hc.fill = PatternFill("solid", fgColor="fee2e2")
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def excel_detay(sonuclar: list[dict], cevap_anahtari: dict, soru_sayisi: int = 20) -> bytes:
    """Detay sayfası: her öğrenci × soru matrisi."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Detay"
    mavi  = PatternFill("solid", fgColor="1a56db")
    beyaz = Font(color="FFFFFF", bold=True)
    kenar = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    basliklar = ["Ad Soyad", "Öğrenci No"] + [f"S{i}" for i in range(1, soru_sayisi + 1)]
    for j, b in enumerate(basliklar, 1):
        h = ws.cell(row=1, column=j, value=b)
        h.fill = mavi
        h.font = beyaz
        h.alignment = Alignment(horizontal="center")
        h.border = kenar
    ws.cell(row=2, column=1, value="CEVAP ANAHTARI").font = Font(bold=True)
    ws.cell(row=2, column=2, value="-")
    for s in range(1, soru_sayisi + 1):
        hc = ws.cell(row=2, column=s + 2, value=cevap_anahtari.get(s, "?"))
        hc.fill = PatternFill("solid", fgColor="dbeafe")
        hc.font = Font(bold=True)
        hc.alignment = Alignment(horizontal="center")
        hc.border = kenar
    for i, s in enumerate(sonuclar, 3):
        ws.cell(row=i, column=1, value=s.get("ad_soyad", "")).border = kenar
        ws.cell(row=i, column=2, value=s.get("ogrenci_no", "")).border = kenar
        for soru in range(1, soru_sayisi + 1):
            c  = s.get("cevaplar", {}).get(soru, "?")
            a  = cevap_anahtari.get(soru, "?")
            hc = ws.cell(row=i, column=soru + 2, value=c)
            hc.alignment = Alignment(horizontal="center")
            hc.border = kenar
            if c == a:
                hc.fill = PatternFill("solid", fgColor="d1fae5")
            elif c == "BOS":
                hc.fill = PatternFill("solid", fgColor="f3f4f6")
            else:
                hc.fill = PatternFill("solid", fgColor="fee2e2")
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 12
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _xls_to_xlsx(xls_bytes: bytes) -> bytes:
    """
    Eski .xls dosyasını .xlsx'e dönüştürür.
    Birçok üniversite sistemi .xls uzantılı HTML dosyası verir —
    bu durumu otomatik algılayıp pandas read_html ile okur.
    """
    import pandas as pd

    # HTML olarak kaydedilmiş .xls dosyasını algıla
    header = xls_bytes[:50].lower()
    if b"<html" in header or b"<?xml" in header or b"<table" in header:
        # HTML tablosunu pandas ile oku
        html_str = xls_bytes.decode("utf-8", errors="ignore")
        tables = pd.read_html(html_str)
        if not tables:
            raise ValueError("HTML dosyasında tablo bulunamadı.")

        # En büyük tabloyu al (genellikle öğrenci listesi)
        df = max(tables, key=len)

        xlsx_wb = Workbook()
        xlsx_ws = xlsx_wb.active
        # Başlıkları yaz
        for col_idx, col_name in enumerate(df.columns, 1):
            xlsx_ws.cell(row=1, column=col_idx, value=col_name)
        # Verileri yaz
        for row_idx, row in enumerate(df.itertuples(index=False), 2):
            for col_idx, val in enumerate(row, 1):
                xlsx_ws.cell(row=row_idx, column=col_idx, value=val)

        buf = BytesIO()
        xlsx_wb.save(buf)
        return buf.getvalue()

    # Gerçek .xls dosyası — xlrd ile oku
    import xlrd

    xls_wb = xlrd.open_workbook(file_contents=xls_bytes)
    xls_ws = xls_wb.sheet_by_index(0)

    xlsx_wb = Workbook()
    xlsx_ws = xlsx_wb.active

    for row_idx in range(xls_ws.nrows):
        for col_idx in range(xls_ws.ncols):
            cell = xls_ws.cell(row_idx, col_idx)
            xlsx_ws.cell(row=row_idx + 1, column=col_idx + 1, value=cell.value)

    buf = BytesIO()
    xlsx_wb.save(buf)
    return buf.getvalue()


def excel_not_girisi(
    sablon_bytes: bytes,
    sonuclar: list[dict],
    not_turu: str = "Vize",
    dosya_adi: str = "",
) -> tuple[bytes, int, int]:
    """
    Üniversite not giriş Excel şablonuna OMR puanlarını yazar.

    Args:
        sablon_bytes: Orijinal Excel dosyası (bytes)
        sonuclar: OMR tarama sonuçları listesi
        not_turu: "Vize" veya "Final" — hangi sütuna yazılacak
        dosya_adi: Orijinal dosya adı (.xls/.xlsx tespiti için)

    Returns:
        (doldurulmuş_excel_bytes, eşleşen_sayı, toplam_öğrenci_sayısı)
    """
    # .xls dosyasını .xlsx'e dönüştür
    if dosya_adi.lower().endswith(".xls") and not dosya_adi.lower().endswith(".xlsx"):
        sablon_bytes = _xls_to_xlsx(sablon_bytes)

    wb = load_workbook(BytesIO(sablon_bytes))
    ws = wb.active

    # ── 1) Başlık satırını bul (Öğrenci Numarası veya Sıra içeren satır) ──
    baslik_satir = None
    no_sutun = None       # Öğrenci Numarası sütunu
    not_sutun = None      # Vize veya Final sütunu

    for row_idx in range(1, min(ws.max_row + 1, 30)):  # İlk 30 satırda ara
        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val and str(val).strip().lower() in ("öğrenci numarası", "ogrenci numarasi"):
                baslik_satir = row_idx
                no_sutun = col_idx
                break
        if baslik_satir:
            break

    if baslik_satir is None:
        # Alternatif: "Sıra" ile başlayan satır bul, öğrenci no bir sonraki sütun
        for row_idx in range(1, min(ws.max_row + 1, 30)):
            for col_idx in range(1, ws.max_column + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val and str(val).strip().lower() in ("sıra", "sira"):
                    baslik_satir = row_idx
                    no_sutun = col_idx + 1  # Öğrenci No, Sıra'nın yanında
                    break
            if baslik_satir:
                break

    if baslik_satir is None or no_sutun is None:
        raise ValueError(
            "Excel dosyasında 'Öğrenci Numarası' veya 'Sıra' başlığı bulunamadı. "
            "Üniversite not giriş formatındaki Excel dosyasını yükleyin."
        )

    # ── 2) Not sütununu bul (Vize / Final) ──
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=baslik_satir, column=col_idx).value
        if val and not_turu.lower() in str(val).strip().lower():
            not_sutun = col_idx
            break

    if not_sutun is None:
        raise ValueError(
            f"Excel dosyasında '{not_turu}' sütunu bulunamadı. "
            f"Başlık satırında (satır {baslik_satir}) '{not_turu}' yazılı sütun olmalı."
        )

    # ── 3) Sonuçları öğrenci numarasına göre indexle ──
    puan_map: dict[str, float] = {}
    for s in sonuclar:
        no = str(s.get("ogrenci_no", "")).strip()
        if no and no != "?" and not s.get("hata"):
            puan_map[no] = s.get("puan", 0)

    # ── 4) Her öğrenci satırında eşleşen numaraya puanı yaz ──
    eslesen = 0
    toplam = 0
    veri_baslangic = baslik_satir + 1

    for row_idx in range(veri_baslangic, ws.max_row + 1):
        no_cell = ws.cell(row=row_idx, column=no_sutun)
        no_val = no_cell.value
        if no_val is None or str(no_val).strip() == "":
            continue  # Boş satır, atla

        toplam += 1
        ogrenci_no = str(int(no_val)) if isinstance(no_val, (int, float)) else str(no_val).strip()

        if ogrenci_no in puan_map:
            not_cell = ws.cell(row=row_idx, column=not_sutun)
            not_cell.value = puan_map[ogrenci_no]
            eslesen += 1

    # ── 5) Kaydet ve döndür ──
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue(), eslesen, toplam
